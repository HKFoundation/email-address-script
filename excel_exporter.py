"""
Excel 输出模块
按日期组织邮件数据并写入 Excel
"""

import os
from datetime import datetime
from typing import List, Dict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Excel 导出器"""

    # 列定义
    COLUMNS = [
        ("序号", 8),
        ("发件邮箱", 35),
        ("邮件标题", 35),
        ("邮件链接", 20),
        ("联系信息", 15),
        ("收件时间", 30),
    ]

    # 表头样式
    HEADER_FONT = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 数据样式
    DATA_FONT = Font(name='微软雅黑', size=10)
    DATA_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ALT_ROW_FILL = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')

    # 边框样式
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 行高设置
    DEFAULT_ROW_HEIGHT = 15  # 默认行高（磅）
    ROW_HEIGHT_INCREMENT = 15  # 行高增量（磅）

    def __init__(self, output_dir: str):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def _create_workbook(self) -> Workbook:
        """创建工作簿并设置样式"""
        wb = Workbook()
        ws = wb.active
        ws.title = "邮件记录"

        # 设置列宽
        for col_idx, (_, width) in enumerate(self.COLUMNS, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 设置行高
        ws.row_dimensions[1].height = 35

        # 写入表头
        for col_idx, (header, _) in enumerate(self.COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

        return wb

    def _apply_row_style(self, ws, row_idx: int, total_rows: int):
        """应用行样式"""
        # 计算每列内容换行后的行数
        max_lines = 1
        for col_idx in range(1, len(self.COLUMNS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value and self.COLUMNS[col_idx - 1][0] != "邮件链接":
                text = str(cell.value).replace(f'=HYPERLINK("{os.path.join(os.getcwd(), "")}', '')
                # 粗略估算换行数（按每行50字符计算）
                lines = max(1, (len(text) + 49) // 50)
                max_lines = max(max_lines, lines)
        
        # 设置行高：基础高度 + (换行数 - 1) * 每行高度
        base_height = self.DEFAULT_ROW_HEIGHT + self.ROW_HEIGHT_INCREMENT
        ws.row_dimensions[row_idx].height = base_height + (max_lines - 1) * self.DEFAULT_ROW_HEIGHT

        for col_idx in range(1, len(self.COLUMNS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = self.DATA_FONT
            cell.alignment = self.DATA_ALIGNMENT
            cell.border = self.THIN_BORDER

            # 交替行背景色
            if row_idx % 2 == 0:
                cell.fill = self.ALT_ROW_FILL

    def _get_filename(self) -> str:
        """生成文件名"""
        return datetime.now().strftime("%Y-%m-%d %H-%M-%S") + ".xlsx"

    def _get_filepath(self, folder_name: str) -> str:
        """生成完整文件路径"""
        filename = self._get_filename()
        return os.path.join(folder_name, filename)

    def export_emails(self, emails: List[Dict]) -> str:
        """将邮件导出到单个 Excel 文件"""
        wb = self._create_workbook()
        ws = wb.active

        # 创建外层文件夹（使用时间戳命名）
        timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
        export_folder = os.path.join(self.output_dir, timestamp)
        os.makedirs(export_folder, exist_ok=True)

        # 创建邮件内容目录（去除下划线）
        mail_dir = os.path.join(export_folder, f"邮件内容 {timestamp}")
        os.makedirs(mail_dir, exist_ok=True)

        # 写入数据
        for idx, email in enumerate(emails, 1):
            row_idx = idx + 1  # 从第2行开始 (第1行是表头)

            # 序号
            ws.cell(row=row_idx, column=1, value=idx)

            # 发件人邮箱
            sender_email = email.get('sender_email', '') or '未知'
            ws.cell(row=row_idx, column=2, value=sender_email)

            # 邮件标题
            subject = email.get('subject', '无主题')
            ws.cell(row=row_idx, column=3, value=subject)

            # 创建邮件正文 txt 文件
            body_text = email.get('body_text', '')
            if body_text:
                txt_filename = f"邮件_{idx}.txt"
                txt_filepath = os.path.join(mail_dir, txt_filename)
                try:
                    with open(txt_filepath, 'w', encoding='utf-8') as f:
                        f.write(f"发件人: {sender_email}\n")
                        f.write(f"主题: {subject}\n")
                        f.write(f"时间: {email.get('received_at', '')[:19].replace('T', ' ')}\n")
                        f.write("=" * 50 + "\n\n")
                        f.write(body_text)
                    link_text = f'=HYPERLINK("{txt_filepath}","查看正文")'
                except Exception:
                    link_text = ''
            else:
                link_text = ''
            ws.cell(row=row_idx, column=4, value=link_text)

            # 联系信息 - 预留列，暂不填充
            ws.cell(row=row_idx, column=5, value='')

            # 收件时间
            received = email.get('received_at', '')
            if isinstance(received, str) and received:
                received_time = received[:19].replace('T', ' ')
                ws.cell(row=row_idx, column=6, value=received_time)
            else:
                ws.cell(row=row_idx, column=6, value='')

            self._apply_row_style(ws, row_idx, len(emails))

        # 保存到外层文件夹
        filepath = self._get_filepath(export_folder)
        wb.save(filepath)

        return filepath

    def export_all(self, emails: List[Dict]) -> Dict[str, str]:
        """
        将所有邮件导出到单个 Excel 文件

        Args:
            emails: 邮件列表

        Returns:
            单个文件路径
        """
        filepath = self.export_emails(emails)
        return {"导出文件": filepath}

    def get_existing_dates(self) -> List[str]:
        """获取已存在的导出文件日期"""
        dates = []
        if os.path.exists(self.output_dir):
            for filename in os.listdir(self.output_dir):
                if filename.endswith('.xlsx') and len(filename) == 13:
                    dates.append(filename.replace('.xlsx', ''))
        return sorted(dates)
